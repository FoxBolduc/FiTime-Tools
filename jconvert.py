#Fox Bolduc (2/12/2017)
#Python 3.6
#Excel Spreadsheet to JSON file converter for FiTime App database
#Use with appropriate jcontemplate.xlsx based file (must be .xlsx)

import sys
from openpyxl import load_workbook

#open the workbook and create .txt file
toLoad = ''
if (len(sys.argv) > 1):
    toLoad = sys.argv[1]+'.xlsx'
else:
    toLoad = 'jcontemplate.xlsx'
wb = load_workbook(toLoad)
ws = wb['Sheet1'] #sheet 1 is the only sheet that will be used with this tool
file_title = ws['B1'].value #designated in template file
file_name = file_title + '.json'
file = open(file_name, "w")

#Locations of dimensional data
num_cols = int(ws['A13'].value)
num_rows = int(ws['B13'].value)

maxCol=1+num_cols
maxRow=14+num_rows

object_keys = []
#convert list
list_name = ws['A14'].value
file.write('{\n\"' + list_name + '\":[\n')

#gets key names 
for row in ws.iter_rows(min_row=14, max_row=14, min_col=2, max_col=maxCol):
    for cell in row:
        object_keys.append(cell.value)

curRow=14
curCol=2

for row in ws.iter_rows(min_row=15, max_row=maxRow-1, min_col=2, max_col=maxCol):
    file.write('{\n')
    i = 0
    for cell in row:
        if (curCol != maxCol):
            file.write('\"' + object_keys[i] + '\":\"' + cell.value + '\",\n')
            curCol = curCol + 1
        else: #gets rid of comma for last key-value pair for an object
            file.write('\"' + object_keys[i] + '\":\"' + cell.value + '\"\n')
            curCol = 2
        i = i+1
    file.write('},\n')

#Changes for last item in list
i = 0
file.write('{\n')
for col in ws.iter_cols(min_row=maxRow, max_row=maxRow, min_col=2, max_col=maxCol):
    for cell in col:
        if (curCol != maxCol):
            file.write('\"' + object_keys[i] + '\":\"' + cell.value + '\",\n')
            curCol = curCol + 1
        else:
            file.write('\"' + object_keys[i] + '\":\"' + cell.value + '\"\n')
        i = i+1
file.write('}]}')

  
    
file.close()
